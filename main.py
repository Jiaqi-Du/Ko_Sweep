import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from sklearn.metrics import mean_squared_error
from openpyxl import load_workbook
from openpyxl.styles import Alignment

'''
Read Input from Dataset.xlsx
'''
data = pd.read_excel('Dataset.xlsx', usecols=['KoD', 'KoA', 'KoV', 'deltaS','tD'])
KoD_list = data['KoD'].tolist()
KoV_list = data['KoV'].tolist()
KoA_list = data['KoA'].tolist()
deltaS_list = data['deltaS'].tolist()
td_max_list = data['tD'].tolist()
nums = len(KoD_list)

# This is time point we have, then we will add time (0,0) and (1,1), there for the intersections is sections + 2 - 1
# Intersections is fixed as 1001
points = 1000
plot_index = 0

# For writing the results to the excel file
wb = load_workbook('Dataset.xlsx')
ws = wb.active

header = {cell.value: cell.column for cell in ws[1]}
ko_col = header['Ko']
s_col = header['S']
mse_col = header['MSE']


if not os.path.exists('Plots'):
    os.makedirs('Plots')


'''
For each row, do Ko simulation
'''

for loop_index in range(nums):
    KoD = KoD_list[loop_index]
    KoA = KoA_list[loop_index]
    KoV = KoV_list[loop_index]
    deltaS = deltaS_list[loop_index]
    td_max = td_max_list[loop_index]
    td = np.linspace(0, td_max, points)
    print(f'Combination {loop_index+1}: KoD = {KoD}, KoA = {KoA}, KoV = {KoV}, S = {deltaS}')

    '''
    I. Caculate displacement efficiency
    '''
    def calculate_RF_td(t):
        if deltaS / KoD < t < deltaS * KoD:
            RF_td = (2 * (KoD * deltaS * t) ** 0.5 - deltaS - t) / (KoD - 1)
        elif t >= deltaS * KoD:
            RF_td = deltaS
        else:
            RF_td = t
        return RF_td
    RF_td = [calculate_RF_td(td) for td in td]

    plot_index += 1
    plt.figure(plot_index)
    plt.plot(td, RF_td, c = 'grey')
    plt.xlabel('tD')
    plt.ylabel('RF')
    plt.title('Displacement Efficiency')
    plt.xlim(0,td_max)
    plt.ylim(0, 1)
    plt.grid(True)
    plt.annotate(f'KoD = {KoD}', xy=(0.8, 0.3), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'S = {deltaS}', xy=(0.8, 0.2), xycoords='axes fraction', fontsize=12, verticalalignment='top')

    #Save the plot
    plt.savefig(f'Plots/{loop_index+1}_displacement.jpg', format='jpg', dpi = 600)
    plt.close()



    '''
    II. Combine Areal and displacement efficiency
    '''
    # Assume the reservoir in horizontal is split into 100 small areas
    C_A = np.sort(np.random.rand(100)) 
    F_A = KoA * C_A / (1 + (KoA-1) * C_A)

    C_Ai = []
    F_Ai = []

    for i in range(len(C_A)):
        if i == 0:
            C_Ai.append(C_A[i] - 0)
            F_Ai.append(F_A[i] - 0)
            continue
        else:
            C_Ai.append(C_A[i] - C_A[i-1])
            F_Ai.append(F_A[i] - F_A[i-1])

    C_Ai.append(1 - C_A[-1])
    F_Ai.append(1 - F_A[-1])

    C_Ai = np.array(C_Ai)
    F_Ai = np.array(F_Ai)
    FC_Ai = F_Ai/C_Ai
    td_A = np.linspace(0, td_max, points)

    RF_CA = []

    for ti in td_A:
        tdi = ti * np.array(FC_Ai)
        RF_CAi = 0
        for i in range (len(FC_Ai)):
            RF_CAi += C_Ai[i] * calculate_RF_td(tdi[i])
        RF_CA.append(RF_CAi)

    RF_CA = np.array(RF_CA)

    # Fit RF_CA vs. td_A, find the best Koval Value and S
    def RF_func(t, K, S):
        return np.piecewise(t, 
                            [t <= S / K, t >= S * K, (S / K < t) & (t < S * K)], 
                            [lambda t: t, lambda t: S , lambda t: (2 * (K * S * t) ** 0.5 - S - t) / (K - 1)])

    popt, _ = curve_fit(RF_func, td_A, RF_CA, bounds=([1, 0], [1000, 1]))
    #print(popt)

    RF_CA_fit_K = popt[0]
    RF_CA_fit_S = popt[1]

    # Numerical solution for RF_CA
    def fit_RF_CA_td(t, K, S):
        if S / K < t < S * K:
            RF_CA_fit = (2 * (K * S * t) ** 0.5 - S - t) / (K - 1)
        elif t >= S * K:
            RF_CA_fit = S
        else:
            RF_CA_fit = t
        return RF_CA_fit


    RF_CA_fit = [fit_RF_CA_td(td, RF_CA_fit_K, RF_CA_fit_S) for td in td_A]
    mse_A = mean_squared_error(RF_CA, RF_CA_fit)

    plot_index += 1
    plt.figure(plot_index)
    # Koval fit
    plt.plot(td_A, RF_func(td_A, *popt), 'b', label='Koval Fit')
    # Sample data
    plt.plot(td_A, RF_CA, 'grey', label='Analytical Solution', markersize=1)
    plt.xlabel('tD')
    plt.ylabel('RF_CA')
    plt.title('Areal Efficiency')
    plt.xlim(0,td_max)
    plt.ylim(0, 1)
    plt.grid(True)
    plt.annotate(f'KoD = {KoD}', xy=(0.75, 0.4), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'KoA = {KoA}', xy=(0.75, 0.3), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'S = {deltaS}', xy=(0.75, 0.2), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'KoDA = {RF_CA_fit_K:.3f}, S* = {RF_CA_fit_S:.3f}', xy=(0.5, 0.1), xycoords='axes fraction', 
                fontsize=12, verticalalignment='top', horizontalalignment='center')    

    plt.legend()
    plt.savefig(f'Plots/{loop_index+1}_Areal.jpg', format='jpg', dpi = 600)
    plt.close()


    '''
    III. Combine Vertical, Areal and Displacement Efficiency
    '''
    # Assume the resevoir is split into 100 layers
    C_V = np.sort(np.random.rand(100))
    F_V = KoV * C_V / (1 + (KoV-1) * C_V)
    #plt.plot(C_V, F_V)

    C_Vi = []
    F_Vi = []

    for i in range(len(C_V)):
        if i == 0:
            C_Vi.append(C_V[i] - 0)
            F_Vi.append(F_V[i] - 0)
            continue
        else:
            C_Vi.append(C_V[i] - C_V[i-1])
            F_Vi.append(F_V[i] - F_V[i-1])

    C_Vi.append(1 - C_V[-1])
    F_Vi.append(1 - F_V[-1])

    C_Vi = np.array(C_Vi)
    F_Vi = np.array(F_Vi) # len(F_Ai) = 101, becasue we add (0,0) and (1,1) to the list
    FC_Vi = F_Vi/C_Vi
    td_V = np.linspace(0, td_max, points)
    RF_CV = []


    for ti in td_V:
        tdi = ti * np.array(FC_Vi)
        RF_CVi = 0
        for i in range (len(FC_Vi)):
            RF_CVi += C_Vi[i] * fit_RF_CA_td(tdi[i], RF_CA_fit_K, RF_CA_fit_S)
        RF_CV.append(RF_CVi)

    RF_CV = np.array(RF_CV)
    #plt.plot(td_A, RF_CV, 'ro')

    popt, _ = curve_fit(RF_func, td_V, RF_CV, bounds=([1, 0], [1000, 1]))
    #print(popt)

    RF_CV_fit_K = popt[0]
    RF_CV_fit_S = popt[1]

    # This is a numerical solution for RF_CV
    def fit_RF_CV_td(t, K, S):
        if S / K < t < S * K:
            RF_CV_fit = (2 * (K * S * t) ** 0.5 - S - t) / (K - 1)
        elif t >= S * K:
            RF_CV_fit = S
        else:
            RF_CV_fit = t
        return RF_CV_fit


    RF_CV_fit = [fit_RF_CV_td(td, RF_CV_fit_K, RF_CV_fit_S) for td in td_V]
    mse = mean_squared_error(RF_CV, RF_CV_fit)
    print(f'Combination {loop_index+1} finished:Ko = {RF_CV_fit_K:.3f}, S* = {RF_CV_fit_S:.3f}, mse = {mse:.3e}')
    
    # Write the values into the Excel file using header names
    ws.cell(row=loop_index+2, column=ko_col, value=f'{RF_CV_fit_K:.3f}')
    ws.cell(row=loop_index+2, column=s_col, value=f'{RF_CV_fit_S:.3f}')
    ws.cell(row=loop_index+2, column=mse_col, value=f'{mse:.3e}')


    plot_index += 1
    plt.figure(plot_index)
    # Koval fit
    plt.plot(td_V, RF_func(td_V, *popt), 'r-', label='Koval Fit')
    # Analytical solution
    plt.plot(td_V, RF_CV, 'grey', label='Analytical Solution', markersize=1)
    plt.xlabel('tD')
    plt.ylabel('RF_T')
    plt.title('Total Efficiency')
    plt.xlim(0,td_max)
    plt.ylim(0, 1)
    plt.grid(True)
    plt.annotate(f'KoD = {KoD}', xy=(0.75, 0.4), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'KoA = {KoA}', xy=(0.75, 0.3), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'KoV = {KoV}', xy=(0.75, 0.2), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'S = {deltaS}', xy=(0.75, 0.1), xycoords='axes fraction', fontsize=12, verticalalignment='top')
    plt.annotate(f'Ko = {RF_CV_fit_K:.3f}, S* = {RF_CV_fit_S:.3f}', xy=(0.5, 0.1), xycoords='axes fraction', 
                fontsize=12, verticalalignment='top', horizontalalignment='center')    

    plt.legend()
    plt.savefig(f'Plots/{loop_index+1}_Total_Efficiency.jpg', format='jpg', dpi = 600)

plt.show()

# Center Alignment
min_col = min(ko_col, s_col, mse_col)
max_col = max(ko_col, s_col, mse_col)
max_row = ws.max_row
for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=min_col, max_col=max_col):
    for cell in r:
        cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save('Dataset.xlsx')

